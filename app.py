import os
import pandas as pd
from flask import Flask, render_template, request, redirect
from openpyxl import load_workbook

app = Flask(__name__)

# Percorsi dei file
original_file = "C:\\Users\\SNAVE\\Desktop\\app turni\\dati.xlsm"
csv_file = "C:\\Users\\SNAVE\\Desktop\\app turni\\dati_modificato.csv"
sheet_name = "Visualizza"

@app.route("/", methods=["GET", "POST"])
def show_data():
    if request.method == "POST":
        try:
            # Ottieni i dati modificati dal form
            edited_data = request.form.to_dict(flat=False)
            print("Dati ricevuti dal form:", edited_data)  # Debug: verifica i dati ricevuti

            # Carica il file Excel e leggiamo solo i valori (senza formule)
            wb = load_workbook(original_file, data_only=True)
            ws = wb[sheet_name]

            # Convertiamo il foglio in un DataFrame
            data = [[cell.value for cell in row] for row in ws.iter_rows()]
            df = pd.DataFrame(data)
            df.columns = df.iloc[0]  # Usa la prima riga come intestazione
            df = df[1:].copy()  # Rimuove la prima riga (header)

            # **Aggiorna i dati modificati in modo corretto**
            for col in df.columns:
                if col in edited_data:
                    df[col] = pd.Series(edited_data[col])  # Usa una Serie Pandas per garantire la modifica corretta

            # **Salviamo solo il foglio "Visualizza" come CSV**
            df.to_csv(csv_file, index=False, encoding="utf-8")

            # **Verifica se il CSV è leggibile**
            df_test = pd.read_csv(csv_file)
            print("Dati salvati nel CSV:", df_test.head())  # Stampa le prime righe per verifica

            return redirect("/")  # Ricarica la pagina dopo aver salvato

        except Exception as e:
            return f"<p style='color:red;'>Errore nel salvataggio del file CSV: {str(e)}</p>"

    df = pd.read_excel(original_file, sheet_name=sheet_name, engine="openpyxl")

    # Creazione della tabella HTML
    table_html = "<form method='POST'><table class='table table-bordered'>"
    table_html += "<tr>" + "".join(f"<th>{col}</th>" for col in df.columns) + "</tr>"

    for index, row in df.iterrows():
        table_html += "<tr>" + "".join(f"<td><input type='text' name='{col}[{index}]' value='{row[col]}'></td>" for col in df.columns) + "</tr>"

    table_html += "</table><button type='submit' class='btn btn-primary mt-3'>Salva Modifiche</button></form>"

    return render_template("index.html", table_data=table_html)

if __name__ == "__main__":
    app.run(debug=True)